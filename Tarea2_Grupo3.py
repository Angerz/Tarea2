import pandas as pd
import numpy as np 
import os.path
import openpyxl

def velocidad_salida(v, n2,n3,n4,n5):
    """Funcion para hallar la velocidad de salida del tren  compuesto de engranajes revertidos

    Args:
        v (float): Velocidad de entrada_en RPM.
        n2 (int): Número de dientes del engranaje 2
        n3 (int): Número de dientes del engranaje 3
        n4 (int): Número de dientes del engranaje 4
        n5 (int): Número de dientes del engranaje 5

    Returns:
        int: Velocidad de salida.
    """
    return v*(n2/n3)*(n4/n5)
#Calcular el diámetrod de paso de cada engranaje
def calcular_diametro_paso(num_dientes, m):
    diametro_paso = num_dientes * m
    return diametro_paso

#Calcular la velocidad de paso de cada engranaje
def velocidad_engranaje(n2,n3,n4,n5):
    w2 = v
    w3 = w2*(n2/n3)
    w4 = w3
    w5 = w4*(n4/n5)
    w_vector = (w2,w3,w4,w5)
    return w_vector

#Calcular la velocidad de linea del engranaje
def velocidad_linea(w,d):
    return (np.pi*d*w)/12

#Calcular las fuerza tangencial del engranaje
def fuerzas_engranajes(p,v):
    return (33000*p)/v

###########################################################################################################3
#INICIO DEL PROGRAMA

if __name__ == "__main__":

    #Se configura el tamaño máximo de la caja como infinito para evitar errores
    mejor_tamaño = float('inf')
    
    #Se crean tuplas y vectores útiles
    mejores_tamaños = ()
    vel_lin = []
    nom = []
    fuerzas_tangenciales=[]
    fuerzas_radiales = []
    Torques = []
    print_f = []
    fuerza_eje = []

    #Se solicita la velocidad de entrada
    v = float(input('Ingresa la velocidad de entrada V en rpm: '))
    print("""
#######################################
#                                     #
#              Inciso 1               #
#                                     #
#######################################
            """)
    
    #Probamos un número de dientes para cada engranaje un un rango de 12 a 100
    for n2 in range(12, 100):
        for n3 in range(12, 100):
            for n4 in range(12, 100):
                for n5 in range(12, 100):
                    #Primera condición: La suma de los diámetros (y por ende el número de dientes) de los engranajes 2 y 3 debe ser igual a la de los engranjes 4 y 5
                    if n2 + n3 == n4 + n5:
                        #Segunda condición: Los engranajes deben ser de diferentes tamaños, además de que 2 y 4 deben ser más pequeños que 3 y 5 respectivamente.
                        if n2 > n4 and n3 != n4 and n3 != n5 and n4 != n5 and n2 != n3:
                            #Si estas condiciones se cumplen, se calcula la velocidad de salida para este conjunto de valores y se guardan en velocidad_actual.
                            velocidad_actual = velocidad_salida(v,n2,n3,n4,n5)
                            #Tercera condición: Se comprueba que la velocidad de salida obtenida esté dentro del rango solicitado.
                            if velocidad_actual >= 150 and velocidad_actual <= 300:
                                #Si esto se cumple, se guarda el tamaño horizontal de la caja en la iteración actual en "tamaño_actual".
                                tamaño_actual = n2+n3
                                #Cuarta condición: Se comparan de manera iterativa todos los tamaños que cumplan las 3 primeras condiciones y se va guardando el menor de todos
                                if tamaño_actual <= mejor_tamaño:
                                    #Si el nuevo tamaño es menor que el anterior, se guardan los datos de esta iteración (números de dientes y velocidad de salida)
                                    mejor_tamaño = tamaño_actual
                                    mejores_tamaños = (n2, n3, n4, n5)
                                    velocidad_optima = velocidad_actual
                            else:
                                pass
                        else:
                            pass
                    else:
                        pass    
    #Prueba si se logró encontrar valores satisfactorios para la velocidad ingresada.
    try:
        print("Los números de dientes adecuados para minimizar el tamaño de la caja son: ")
        for i in range(4):
            print("Engranaje", i+2,": ", mejores_tamaños[i])
        
        print(f"Con una velocidad de salida de {round(velocidad_optima,3)} rpm.")
    #De no encontrarse suelta una excepción
    except:
        print("No se encotraron valores para el número de dientes de los engranajes para la velocidad ingresada")

    print("""
#######################################
#                                     #
#              Inciso 2               #
#                                     #
#######################################
            """)
    
    #Solicitamos otros parámetros necesarios
    hp = float(input("Potencia requerida para transmitir en hp: "))
    m = float(input("Módulo de engrane en mm: "))
    m = 0.0393701*m
    #Cálculo de diámetros para cada engranaje
    diametro_paso_entrada = calcular_diametro_paso(mejores_tamaños[0], m)
    diametro_paso_intermedio1 = calcular_diametro_paso(mejores_tamaños[1], m)
    diametro_paso_intermedio2 = calcular_diametro_paso(mejores_tamaños[2], m)
    diametro_paso_salida = calcular_diametro_paso(mejores_tamaños[3], m)
    #Se crea un vector con estos valores
    d_vector = [diametro_paso_entrada,diametro_paso_intermedio1,diametro_paso_intermedio2,diametro_paso_salida]

    #Se crea un vector con las velocidades angulares de cada engranaje
    w_vector = velocidad_engranaje(mejores_tamaños[0],mejores_tamaños[1],mejores_tamaños[2],mejores_tamaños[3])

    #Se calculan las velocidades de línea de cada engranaje
    for i in range(4):
        vel_lineal = velocidad_linea(w_vector[i],d_vector[i])
        vel_lin.append(vel_lineal)

    #Se calculan las fuerzas y torques para cada
    for i in range(4):
        f = fuerzas_engranajes(hp,vel_lin[i])
        fuerzas_tangenciales.append(f)
        Torques.append(f*d_vector[i]/2)
        fuerzas_radiales.append(f*np.tan(20*(np.pi/180)))

    #Se crea un vector con las fuerzas radiales y tangenciales de cada engranaje
    for i in range(4):
        print_f.append(fuerzas_radiales[i])
        print_f.append(fuerzas_tangenciales[i])

    print("Fuerzas y Torques ejercidas en los ejes")

    #Imprime las fuerzas en los ejes
    for i in range(4):
        resul = np.sqrt(fuerzas_radiales[i]**2 + fuerzas_tangenciales[i]**2)
        print("Fuerza que ejerce el engrane ", i+2,"en el eje:",resul)
        fuerza_eje.append(resul)
    #Imprime los torques de entrada y salida
    print("Torque de entrada", Torques[0])
    print("Torque de salida", Torques[3])

    print("""
#######################################
#                                     #
#              Inciso 3               #
#                                     #
#######################################
            """)
    
    #Imprime los diámetros de cada engranajes
    for i in range(4):
        print("El diámetro de paso de el engranaje ",i+2, " es: ",d_vector[i]," pulgadas.")

    print("""
#######################################
#                                     #
#              Inciso 4               #
#                                     #
#######################################
            """)
    
    #Imprime las velocidades de línea de cada engranaje
    for i in range(4):
        print("La velocidad de línea del engranaje ", i+2, "es: ", vel_lin[i], "pies/min")

    
    print("""
#######################################
#                                     #
#              Inciso 5               #
#                                     #
#######################################
            """)
    #Arreglo útil para imprimir las fuerzas
    nombre2 = ["Fuerza radial transmitida de 2 a 3: ", "Fuerza tangencial transmitida de 2 a 3: ", "Fuerza radial transmitida de 3 a 2: ", 
              "Fuerza tangencial transmitida de 3 a 2: ","Fuerza radial transmitida de 4 a 5: ", "Fuerza tangencial transmitida de 4 a 5: ", 
              "Fuerza radial transmitida de 5 a 4: ", "Fuerza tangencial transmitida de 5 a 4: " ]
    #Se imprimen las fuerzas tangenciales y radiales
    for i in range(8):
        print(nombre2[i], print_f[i]," lbf.")

    print("NO EXISTEN FUERZAS AXIALES TRANSMITIDAS PARA ENGRANAJES RECTOS")

    ##################################################################################################################################
    #GUARDANDO LOS RESULTAODS EN EXCEL

    #Nombre del archivo
    nombre_archivo = "resultados.xlsx"

    #Agregamos todos los datos obtenidos que sean relevantes al array "datos_engranajes"
    datos_engranjes = []
    for i in range(4):
        dientes_engranaje = mejores_tamaños[i]
        fuerza_sobre_eje = fuerza_eje[i]
        diametro_paso = d_vector[i]
        velocidades_linea = vel_lin[i]
        fuerzas_tangenciales_engranajes = fuerzas_tangenciales[i]
        fuerzas_radiales_engranajes = fuerzas_radiales[i]
        datos_engranjes.append([dientes_engranaje, fuerza_sobre_eje, diametro_paso, velocidades_linea, fuerzas_tangenciales_engranajes, fuerzas_radiales_engranajes])

    # Nombres de las columnas y filas
    nombres_columnas = ["Numero de Dientes", "Fuerza sobre el eje" ,"Diametro de Paso", "Velocidad Linea", "Fuerza tangencial", "Fuerza radial"]
    nombres_filas = [f"Engranaje {i+2}" for i in range(4)] #Uso de list comprehension para reducir líneas.

    # Crea un DataFrame con los datos y los nombres de filas y columnas
    df = pd.DataFrame(datos_engranjes, columns=nombres_columnas, index=nombres_filas)

    # Verifica si el archivo de Excel existe
    if os.path.isfile(nombre_archivo):
        # Si el archivo existe, carga el libro de trabajo existente y agrega una nueva hoja
        book = openpyxl.load_workbook(nombre_archivo)
        writer = pd.ExcelWriter(nombre_archivo, engine='openpyxl')
        writer.book = book
        
        #El nombre de cada hoja se titula igual que la velocidad ingresada
        hoja_nombre = f"Velocidad = {v}"
        
        # Agrega el DataFrame a la nueva hoja
        df.to_excel(writer, sheet_name=hoja_nombre)

        #Trae la hojas actual para escribir en ella
        sheet = book[hoja_nombre]
        
        # Agrega celdas adicionales con los datos de potencia y módulo proporcionadas por el usuario
        fila_ultima = len(df) + 1
        sheet.cell(row=fila_ultima + 1, column=1, value="Potencia de entrada en HP")
        sheet.cell(row=fila_ultima + 1, column=2, value=hp)
        sheet.cell(row=fila_ultima + 2, column=1, value="Módulo en pulgadas")
        sheet.cell(row=fila_ultima + 2, column=2, value=m)
    else:
        # Si el archivo no existe, crea un nuevo libro de trabajo y guarda el DataFrame en la primera hoja
        writer = pd.ExcelWriter(nombre_archivo, engine='openpyxl')
        df.to_excel(writer, sheet_name=f"Velocidad = {v}")

        #Jala la hoja actual
        book = writer.book
        sheet = book[f"Velocidad = {v}"]
        
        # Agrega celdas adicionales con los datos de potencia y módulo proporcionadas por el usuario
        fila_ultima = len(df) + 1
        sheet.cell(row=fila_ultima + 1, column=1, value="Potencia de entrada en HP")
        sheet.cell(row=fila_ultima + 1, column=2, value=hp)
        sheet.cell(row=fila_ultima + 2, column=1, value="Módulo en pulgadas")
        sheet.cell(row=fila_ultima + 2, column=2, value=m)

    # Guarda el archivo de Excel
    writer.save()
    writer.close()
